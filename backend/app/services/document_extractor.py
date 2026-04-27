"""Deterministic document extractor — no LLM involved.

Takes raw file bytes and produces a structured representation that the
AI mapper can consume. Supports:
  - Excel (.xlsx, .xlsm, .xls)
  - PDF (text-based; scanned PDFs are flagged, not OCR'd)
  - CSV / TSV

Limits:
  - 50 MB per file
  - 50 sheets per Excel
  - 200 pages per PDF
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any, Optional

from app.core.logging import get_logger

log = get_logger("app.extractor")

# ---------------------------------------------------------------------------
# Size limits
# ---------------------------------------------------------------------------

MAX_FILE_BYTES = 50 * 1024 * 1024   # 50 MB
MAX_EXCEL_SHEETS = 50
MAX_PDF_PAGES = 200


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------

class ExtractionError(Exception):
    """Raised when a document cannot be processed."""
    http_status: int = 400


class FileTooLargeError(ExtractionError):
    http_status: int = 413

    def __init__(self, size_mb: float) -> None:
        super().__init__(f"File too large ({size_mb:.1f} MB). Maximum is {MAX_FILE_BYTES / 1024 / 1024:.0f} MB.")


class UnsupportedFormatError(ExtractionError):
    def __init__(self, mime: str) -> None:
        super().__init__(f"Unsupported file format: {mime}. Upload an Excel (.xlsx), PDF, or CSV file.")


class ScannedPDFError(ExtractionError):
    def __init__(self) -> None:
        super().__init__(
            "This PDF appears to be a scanned image without selectable text. "
            "OCR is not supported yet. Please upload a text-based PDF or Excel file."
        )


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SheetData:
    """Extracted data from a single sheet/tab."""
    name: str
    rows: list[list[Any]]          # raw rows × columns (str | float | int | None)
    row_count: int = 0
    col_count: int = 0
    has_merged_cells: bool = False
    merged_cell_ratio: float = 0.0  # fraction of header rows with merges


@dataclass
class ExtractedDocument:
    """Complete extraction result."""
    file_type: str                           # "excel" | "pdf" | "csv"
    original_filename: str
    file_size_bytes: int
    sheets: list[SheetData] = field(default_factory=list)
    page_count: Optional[int] = None         # for PDFs
    needs_ocr: bool = False
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# MIME detection (by magic bytes, not extension)
# ---------------------------------------------------------------------------

def _detect_mime(data: bytes, filename: str) -> str:
    """Detect file type from magic bytes."""
    if data[:4] == b"PK\x03\x04":
        # ZIP-based: xlsx, xlsm, docx, etc.
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xlsm")):
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # Default to xlsx for unknown ZIP files
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if data[:5] == b"%PDF-":
        return "application/pdf"
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "application/vnd.ms-excel"  # .xls (OLE2)
    # Try CSV/TSV heuristic: first 2000 bytes are ASCII/UTF-8 text
    try:
        sample = data[:2000].decode("utf-8")
        if "\t" in sample or "," in sample:
            return "text/csv"
    except UnicodeDecodeError:
        pass
    return "application/octet-stream"


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------

def _extract_excel(data: bytes, filename: str) -> ExtractedDocument:
    """Extract all sheets from an Excel workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets: list[SheetData] = []

    if len(wb.sheetnames) > MAX_EXCEL_SHEETS:
        raise ExtractionError(f"Too many sheets ({len(wb.sheetnames)}). Maximum is {MAX_EXCEL_SHEETS}.")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows():
            row_data: list[Any] = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append(None)
                elif isinstance(val, (int, float)):
                    row_data.append(val)
                else:
                    row_data.append(str(val).strip())
            rows.append(row_data)

        # Detect merged cells (need non-read-only workbook for this)
        has_merges = False
        merge_ratio = 0.0

        col_count = max((len(r) for r in rows), default=0)
        sd = SheetData(
            name=sheet_name,
            rows=rows,
            row_count=len(rows),
            col_count=col_count,
            has_merged_cells=has_merges,
            merged_cell_ratio=merge_ratio,
        )
        sheets.append(sd)

    wb.close()

    # Re-open in non-read-only mode to detect merged cells
    try:
        wb2 = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=False)
        for sd in sheets:
            ws2 = wb2[sd.name]
            merged = list(ws2.merged_cells.ranges)
            if merged:
                sd.has_merged_cells = True
                # Count merges in first 5 rows (header area)
                header_merges = sum(
                    1 for m in merged if m.min_row <= 5
                )
                header_cells = sd.col_count * min(5, sd.row_count)
                if header_cells > 0:
                    sd.merged_cell_ratio = header_merges / header_cells
        wb2.close()
    except Exception:
        pass  # Non-critical; merged cell detection is best-effort

    return ExtractedDocument(
        file_type="excel",
        original_filename=filename,
        file_size_bytes=len(data),
        sheets=sheets,
    )


# ---------------------------------------------------------------------------
# PDF extraction
# ---------------------------------------------------------------------------

def _extract_pdf(data: bytes, filename: str) -> ExtractedDocument:
    """Extract text and tables from a PDF."""
    import pdfplumber

    pdf = pdfplumber.open(io.BytesIO(data))
    page_count = len(pdf.pages)

    if page_count > MAX_PDF_PAGES:
        pdf.close()
        raise ExtractionError(f"PDF has {page_count} pages. Maximum is {MAX_PDF_PAGES}.")

    all_text = []
    all_tables: list[list[list[Any]]] = []
    total_chars = 0

    for page in pdf.pages:
        text = page.extract_text() or ""
        total_chars += len(text.strip())
        all_text.append(text)

        tables = page.extract_tables()
        if tables:
            for table in tables:
                cleaned: list[list[Any]] = []
                for row in table:
                    cleaned.append([
                        cell.strip() if isinstance(cell, str) else cell
                        for cell in row
                    ])
                all_tables.append(cleaned)

    pdf.close()

    # Detect scanned PDFs (very little text)
    needs_ocr = total_chars < 100 and page_count > 0

    # Build sheets: one sheet from tables, one from raw text
    sheets: list[SheetData] = []
    warnings: list[str] = []

    if all_tables:
        for i, table in enumerate(all_tables):
            col_count = max((len(r) for r in table), default=0)
            sheets.append(SheetData(
                name=f"Table_{i + 1}",
                rows=table,
                row_count=len(table),
                col_count=col_count,
            ))
    elif not needs_ocr:
        # Try to parse text as a table (lines → rows, whitespace → columns)
        for i, text in enumerate(all_text):
            if not text.strip():
                continue
            lines = text.strip().split("\n")
            rows: list[list[Any]] = []
            for line in lines:
                # Split on multiple spaces (common in financial statements)
                parts = [p.strip() for p in line.split("  ") if p.strip()]
                if parts:
                    rows.append(parts)
            if rows:
                col_count = max(len(r) for r in rows)
                sheets.append(SheetData(
                    name=f"Page_{i + 1}",
                    rows=rows,
                    row_count=len(rows),
                    col_count=col_count,
                ))

    if needs_ocr:
        warnings.append("PDF appears to be scanned. OCR is not supported in v1.")

    return ExtractedDocument(
        file_type="pdf",
        original_filename=filename,
        file_size_bytes=len(data),
        sheets=sheets,
        page_count=page_count,
        needs_ocr=needs_ocr,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# CSV extraction
# ---------------------------------------------------------------------------

def _extract_csv(data: bytes, filename: str) -> ExtractedDocument:
    """Extract rows from a CSV or TSV file."""
    text = data.decode("utf-8-sig")  # Handle BOM
    # Detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4000])
    except csv.Error:
        dialect = csv.excel  # fallback to comma

    reader = csv.reader(io.StringIO(text), dialect)
    rows: list[list[Any]] = []
    for row in reader:
        parsed: list[Any] = []
        for cell in row:
            cell = cell.strip()
            if not cell:
                parsed.append(None)
                continue
            # Try to parse as number
            try:
                # Remove thousands separators
                clean = cell.replace(",", "").replace(" ", "")
                parsed.append(float(clean))
            except ValueError:
                parsed.append(cell)
        rows.append(parsed)

    col_count = max((len(r) for r in rows), default=0)
    sheets = [SheetData(
        name="Data",
        rows=rows,
        row_count=len(rows),
        col_count=col_count,
    )]

    return ExtractedDocument(
        file_type="csv",
        original_filename=filename,
        file_size_bytes=len(data),
        sheets=sheets,
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def extract_document(data: bytes, filename: str) -> ExtractedDocument:
    """Extract structured data from a file.

    Args:
        data: Raw file bytes.
        filename: Original filename (used for MIME fallback and display).

    Returns:
        ExtractedDocument with sheets, rows, and metadata.

    Raises:
        FileTooLargeError: File exceeds 50 MB.
        UnsupportedFormatError: Unrecognized file type.
        ScannedPDFError: PDF with no text (needs OCR).
        ExtractionError: Other extraction failures.
    """
    if len(data) > MAX_FILE_BYTES:
        raise FileTooLargeError(len(data) / (1024 * 1024))

    mime = _detect_mime(data, filename)
    log.info("extracting_document", filename=filename, mime=mime, size_bytes=len(data))

    if mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        doc = _extract_excel(data, filename)
    elif mime == "application/pdf":
        doc = _extract_pdf(data, filename)
        if doc.needs_ocr:
            raise ScannedPDFError()
    elif mime == "text/csv":
        doc = _extract_csv(data, filename)
    else:
        raise UnsupportedFormatError(mime)

    # Strip empty sheets
    doc.sheets = [s for s in doc.sheets if s.row_count > 0]

    log.info(
        "extraction_complete",
        filename=filename,
        file_type=doc.file_type,
        sheet_count=len(doc.sheets),
        total_rows=sum(s.row_count for s in doc.sheets),
    )
    return doc
