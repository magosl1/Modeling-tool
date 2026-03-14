"""
Generates Excel templates (historical data + module-specific) using openpyxl.
Templates include embedded formula-based validation checks so users can see
errors while filling in the data, before uploading.
"""
from io import BytesIO
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


PNL_ITEMS = [
    ("Revenue", "(+)"),
    ("Cost of Goods Sold", "(-)"),
    ("Gross Profit", "(+)"),
    ("SG&A", "(-)"),
    ("R&D", "(-)"),
    ("D&A", "(-)"),
    ("Amortization of Intangibles", "(-)"),
    ("Other OpEx", "(-)"),
    ("EBIT", "(+/-)"),
    ("Interest Income", "(+)"),
    ("Interest Expense", "(-)"),
    ("Other Non-Operating Income / (Expense)", "(+/-)"),
    ("EBT", "(+/-)"),
    ("Tax", "(-)"),
    ("Net Income", "(+/-)"),
]

BS_ITEMS = [
    # Fixed Assets
    ("PP&E Gross", "Fixed Assets", "(+)"),
    ("Accumulated Depreciation", "Fixed Assets", "(-)"),
    ("Net PP&E", "Fixed Assets", "(+)"),
    ("Intangibles Gross", "Fixed Assets", "(+)"),
    ("Accumulated Amortization", "Fixed Assets", "(-)"),
    ("Net Intangibles", "Fixed Assets", "(+)"),
    ("Goodwill", "Fixed Assets", "(+)"),
    # Current Assets
    ("Inventories", "Current Assets", "(+)"),
    ("Accounts Receivable", "Current Assets", "(+)"),
    ("Prepaid Expenses & Other Current Assets", "Current Assets", "(+)"),
    # Cash
    ("Cash & Equivalents", "Cash", "(+)"),
    # Non-Operating
    ("Non-Operating Assets", "Non-Operating", "(+)"),
    # Equity
    ("Share Capital", "Equity", "(-)"),
    ("Retained Earnings", "Equity", "(-)"),
    ("Other Equity (AOCI, Treasury Stock, etc.)", "Equity", "(-)"),
    # Current Liabilities
    ("Accounts Payable", "Current Liabilities", "(-)"),
    ("Accrued Liabilities", "Current Liabilities", "(-)"),
    ("Other Current Liabilities", "Current Liabilities", "(-)"),
    # Other Long-Term Liabilities
    ("Other Long-Term Liabilities", "Other Long-Term", "(-)"),
    # Debt
    ("Short-Term Debt", "Debt", "(-)"),
    ("Long-Term Debt", "Debt", "(-)"),
]

CF_ITEMS = [
    ("Net Income", "(+/-)"),
    ("D&A Add-back", "(+)"),
    ("Amortization of Intangibles Add-back", "(+)"),
    ("Changes in Working Capital", "(+/-)"),
    ("Operating Cash Flow", "(+/-)"),
    ("Capex", "(-)"),
    ("Acquisitions / Disposals", "(-)"),
    ("Investing Cash Flow", "(+/-)"),
    ("Debt Issuance / Repayment", "(+/-)"),
    ("Dividends Paid", "(-)"),
    ("Share Issuance / Buyback", "(+/-)"),
    ("Financing Cash Flow", "(+/-)"),
    ("Net Change in Cash", "(+/-)"),
]

# Styles for validation rows
CHECK_FONT = Font(bold=True, size=10)
CHECK_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHECK_LABEL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)


def _style_header(ws, col_count: int):
    """Style the header row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _add_check_row(ws, row: int, label: str, formula_template: str,
                   year_cols: list, col_offset: int):
    """Add a validation check row with per-year IF formulas.
    
    formula_template uses {col} as placeholder for the year column letter.
    Returns the row number used.
    """
    ws.cell(row=row, column=1, value=f"✓ CHECK: {label}")
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True, size=9, color="1F4E79")
    ws.cell(row=row, column=1).fill = CHECK_LABEL_FILL

    # Fill label columns
    for c in range(2, col_offset + 1):
        ws.cell(row=row, column=c).fill = CHECK_LABEL_FILL

    for i, year_col in enumerate(year_cols):
        col_letter = get_column_letter(year_col)
        formula = formula_template.format(col=col_letter, row=row)
        cell = ws.cell(row=row, column=year_col, value=formula)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = CHECK_LABEL_FILL

    return row


def _item_row(item_name: str, items_list, data_start_row: int = 2) -> int:
    """Get the Excel row number for a given item name (1-indexed)."""
    for idx, item in enumerate(items_list):
        name = item[0] if isinstance(item, tuple) else item
        if name == item_name:
            return data_start_row + idx
    return data_start_row  # fallback


def _write_sheet_with_checks(ws, items, years: List[int], units_label: str,
                              include_bucket: bool = False, sheet_type: str = "PNL"):
    """Write a financial data sheet with embedded validation formulas."""
    # Build headers
    if include_bucket:
        headers = ["Line Item", "Bucket", "Sign", "Units"] + [str(y) for y in years]
        data_col_offset = 5  # year data starts at column 5
    else:
        headers = ["Line Item", "Sign", "Units"] + [str(y) for y in years]
        data_col_offset = 4  # year data starts at column 4

    # Write header row
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header(ws, len(headers))

    # Write data rows
    data_start_row = 2
    if include_bucket:
        for row_idx, (item, bucket, sign) in enumerate(items, data_start_row):
            ws.cell(row=row_idx, column=1, value=item)
            ws.cell(row=row_idx, column=2, value=bucket)
            ws.cell(row=row_idx, column=3, value=sign)
            ws.cell(row=row_idx, column=4, value=units_label)
    else:
        for row_idx, (item, sign) in enumerate(items, data_start_row):
            ws.cell(row=row_idx, column=1, value=item)
            ws.cell(row=row_idx, column=2, value=sign)
            ws.cell(row=row_idx, column=3, value=units_label)

    year_cols = list(range(data_col_offset, data_col_offset + len(years)))
    check_start = data_start_row + len(items) + 1  # leave a blank row

    # === P&L checks ===
    if sheet_type == "PNL":
        # Row references (1-indexed)
        r = lambda name: _item_row(name, items, data_start_row)

        # Check 1: Gross Profit = Revenue + COGS
        rev_r, cogs_r, gp_r = r("Revenue"), r("Cost of Goods Sold"), r("Gross Profit")
        formula = f'=IF(ABS({{col}}{gp_r}-({{col}}{rev_r}+{{col}}{cogs_r}))<0.5,"✅ OK","❌ GP ≠ Rev+COGS")'
        _add_check_row(ws, check_start, "Gross Profit = Revenue + COGS", formula, year_cols, data_col_offset - 1)

        # Check 2: EBIT = GP + SGA + R&D + D&A + Amort + Other OpEx
        sga_r = r("SG&A")
        rd_r = r("R&D")
        da_r = r("D&A")
        amort_r = r("Amortization of Intangibles")
        oopex_r = r("Other OpEx")
        ebit_r = r("EBIT")
        formula = (f'=IF(ABS({{col}}{ebit_r}-({{col}}{gp_r}+{{col}}{sga_r}+{{col}}{rd_r}'
                   f'+{{col}}{da_r}+{{col}}{amort_r}+{{col}}{oopex_r}))<0.5,'
                   f'"✅ OK","❌ EBIT ≠ GP+OpEx")')
        _add_check_row(ws, check_start + 1, "EBIT = Gross Profit + OpEx items", formula, year_cols, data_col_offset - 1)

        # Check 3: EBT = EBIT + II + IE + Other Non-Op
        ii_r = r("Interest Income")
        ie_r = r("Interest Expense")
        ono_r = r("Other Non-Operating Income / (Expense)")
        ebt_r = r("EBT")
        formula = (f'=IF(ABS({{col}}{ebt_r}-({{col}}{ebit_r}+{{col}}{ii_r}+{{col}}{ie_r}'
                   f'+{{col}}{ono_r}))<0.5,'
                   f'"✅ OK","❌ EBT ≠ EBIT±NonOp")')
        _add_check_row(ws, check_start + 2, "EBT = EBIT + Interest + NonOp", formula, year_cols, data_col_offset - 1)

        # Check 4: Net Income = EBT + Tax
        tax_r = r("Tax")
        ni_r = r("Net Income")
        formula = f'=IF(ABS({{col}}{ni_r}-({{col}}{ebt_r}+{{col}}{tax_r}))<0.5,"✅ OK","❌ NI ≠ EBT+Tax")'
        _add_check_row(ws, check_start + 3, "Net Income = EBT + Tax", formula, year_cols, data_col_offset - 1)

    # === Balance Sheet checks ===
    elif sheet_type == "BS":
        r = lambda name: _item_row(name, items, data_start_row)

        # Check 1: Net PP&E = PP&E Gross + Acc Dep
        ppeg_r = r("PP&E Gross")
        accd_r = r("Accumulated Depreciation")
        nppe_r = r("Net PP&E")
        formula = f'=IF(ABS({{col}}{nppe_r}-({{col}}{ppeg_r}+{{col}}{accd_r}))<0.5,"✅ OK","❌ Net PPE error")'
        _add_check_row(ws, check_start, "Net PP&E = Gross + Acc Dep", formula, year_cols, data_col_offset - 1)

        # Check 2: Net Intangibles = Intangibles Gross + Acc Amort
        intg_r = r("Intangibles Gross")
        acca_r = r("Accumulated Amortization")
        nint_r = r("Net Intangibles")
        formula = f'=IF(ABS({{col}}{nint_r}-({{col}}{intg_r}+{{col}}{acca_r}))<0.5,"✅ OK","❌ Net Intang error")'
        _add_check_row(ws, check_start + 1, "Net Intangibles = Gross + Acc Amort", formula, year_cols, data_col_offset - 1)

        # Check 3: Assets + Liabilities + Equity = 0
        gw_r = r("Goodwill")
        inv_r = r("Inventories")
        ar_r = r("Accounts Receivable")
        pre_r = r("Prepaid Expenses & Other Current Assets")
        cash_r = r("Cash & Equivalents")
        noa_r = r("Non-Operating Assets")
        sc_r = r("Share Capital")
        re_r = r("Retained Earnings")
        oe_r = r("Other Equity (AOCI, Treasury Stock, etc.)")
        ap_r = r("Accounts Payable")
        acl_r = r("Accrued Liabilities")
        ocl_r = r("Other Current Liabilities")
        olt_r = r("Other Long-Term Liabilities")
        std_r = r("Short-Term Debt")
        ltd_r = r("Long-Term Debt")

        # Total Assets
        assets_sum = (f"{{col}}{nppe_r}+{{col}}{nint_r}+{{col}}{gw_r}"
                      f"+{{col}}{inv_r}+{{col}}{ar_r}+{{col}}{pre_r}"
                      f"+{{col}}{cash_r}+{{col}}{noa_r}")
        # Total L+E
        le_sum = (f"{{col}}{sc_r}+{{col}}{re_r}+{{col}}{oe_r}"
                  f"+{{col}}{ap_r}+{{col}}{acl_r}+{{col}}{ocl_r}+{{col}}{olt_r}"
                  f"+{{col}}{std_r}+{{col}}{ltd_r}")

        formula = f'=IF(ABS(({assets_sum})+({le_sum}))<0.5,"✅ BS Balances","❌ BS DOES NOT BALANCE")'
        _add_check_row(ws, check_start + 2, "Assets + L&E = 0 (Balance Check)", formula, year_cols, data_col_offset - 1)

        # Subtotal rows for convenience
        check_start += 3
        ws.cell(row=check_start, column=1, value="  → Total Assets")
        ws.cell(row=check_start, column=1).font = Font(italic=True, size=9, color="4472C4")
        for yc in year_cols:
            cl = get_column_letter(yc)
            ws.cell(row=check_start, column=yc,
                    value=f"={cl}{nppe_r}+{cl}{nint_r}+{cl}{gw_r}+{cl}{inv_r}+{cl}{ar_r}+{cl}{pre_r}+{cl}{cash_r}+{cl}{noa_r}")
            ws.cell(row=check_start, column=yc).font = Font(italic=True, size=9)

        check_start += 1
        ws.cell(row=check_start, column=1, value="  → Total L & E")
        ws.cell(row=check_start, column=1).font = Font(italic=True, size=9, color="4472C4")
        for yc in year_cols:
            cl = get_column_letter(yc)
            ws.cell(row=check_start, column=yc,
                    value=f"={cl}{sc_r}+{cl}{re_r}+{cl}{oe_r}+{cl}{ap_r}+{cl}{acl_r}+{cl}{ocl_r}+{cl}{olt_r}+{cl}{std_r}+{cl}{ltd_r}")
            ws.cell(row=check_start, column=yc).font = Font(italic=True, size=9)

    # === Cash Flow checks ===
    elif sheet_type == "CF":
        r = lambda name: _item_row(name, items, data_start_row)

        # Check 1: OCF = NI + D&A + Amort + WC
        ni_r = r("Net Income")
        da_r = r("D&A Add-back")
        am_r = r("Amortization of Intangibles Add-back")
        wc_r = r("Changes in Working Capital")
        ocf_r = r("Operating Cash Flow")
        formula = (f'=IF(ABS({{col}}{ocf_r}-({{col}}{ni_r}+{{col}}{da_r}'
                   f'+{{col}}{am_r}+{{col}}{wc_r}))<0.5,"✅ OK","❌ OCF error")')
        _add_check_row(ws, check_start, "OCF = NI + D&A + Amort + WC", formula, year_cols, data_col_offset - 1)

        # Check 2: ICF = Capex + Acq/Disp
        cap_r = r("Capex")
        acq_r = r("Acquisitions / Disposals")
        icf_r = r("Investing Cash Flow")
        formula = f'=IF(ABS({{col}}{icf_r}-({{col}}{cap_r}+{{col}}{acq_r}))<0.5,"✅ OK","❌ ICF error")'
        _add_check_row(ws, check_start + 1, "ICF = Capex + Acq/Disp", formula, year_cols, data_col_offset - 1)

        # Check 3: FCF = Debt + Div + Shares
        dbt_r = r("Debt Issuance / Repayment")
        div_r = r("Dividends Paid")
        shr_r = r("Share Issuance / Buyback")
        fcf_r = r("Financing Cash Flow")
        formula = (f'=IF(ABS({{col}}{fcf_r}-({{col}}{dbt_r}+{{col}}{div_r}'
                   f'+{{col}}{shr_r}))<0.5,"✅ OK","❌ FCF error")')
        _add_check_row(ws, check_start + 2, "FCF = Debt + Div + Shares", formula, year_cols, data_col_offset - 1)

        # Check 4: Net Change = OCF + ICF + FCF
        ncc_r = r("Net Change in Cash")
        formula = (f'=IF(ABS({{col}}{ncc_r}-({{col}}{ocf_r}+{{col}}{icf_r}'
                   f'+{{col}}{fcf_r}))<0.5,"✅ OK","❌ Net ≠ OCF+ICF+FCF")')
        _add_check_row(ws, check_start + 3, "Net Change = OCF + ICF + FCF", formula, year_cols, data_col_offset - 1)


def generate_historical_template(years: List[int], currency: str, scale: str) -> bytes:
    """Generate the 3-tab historical data Excel template with validation checks."""
    units_label = f"{currency} {scale}"
    wb = openpyxl.Workbook()

    # Tab 1 — P&L
    ws_pnl = wb.active
    ws_pnl.title = "P&L"
    _write_sheet_with_checks(ws_pnl, PNL_ITEMS, years, units_label, sheet_type="PNL")

    # Tab 2 — Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    _write_sheet_with_checks(ws_bs, BS_ITEMS, years, units_label, include_bucket=True, sheet_type="BS")

    # Tab 3 — Cash Flow
    ws_cf = wb.create_sheet("Cash Flow")
    _write_sheet_with_checks(ws_cf, CF_ITEMS, years, units_label, sheet_type="CF")

    # Column widths
    for ws in [ws_pnl, ws_bs, ws_cf]:
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        if ws == ws_bs:
            ws.column_dimensions["D"].width = 15
        for i, _ in enumerate(years):
            col_letter = get_column_letter(4 + i if ws != ws_bs else 5 + i)
            ws.column_dimensions[col_letter].width = 18

    # Freeze panes
    for ws in [ws_pnl, ws_cf]:
        ws.freeze_panes = "D2"
    ws_bs.freeze_panes = "E2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_module_template(
    module: str,
    line_items: List[str],
    years: List[int],
    currency: str,
    scale: str,
) -> bytes:
    """Generate a dynamic module template based on user configuration."""
    units_label = f"{currency} {scale}"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = module.upper()

    headers = ["Line Item", "Units"] + [str(y) for y in years]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(line_items, 2):
        ws.cell(row=row_idx, column=1, value=item)
        ws.cell(row=row_idx, column=2, value=units_label)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    for i, _ in enumerate(years):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
