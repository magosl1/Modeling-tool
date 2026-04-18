"""
Validates uploaded historical data against the 8 validation rules specified.
Returns a list of ValidationError objects.

Revenue flexibility: accepts either a single "Revenue" line OR multiple
sub-lines (stored with bucket="Revenue") summing to a "Total Revenue" row.
The validator uses whichever is present for formula cross-checks.
"""
from dataclasses import dataclass
from decimal import Decimal
from typing import Dict, List, Optional

TOLERANCE = Decimal("0.5")

# Standard PNL items that are NOT revenue sub-lines.
# Any other item found before "Cost of Goods Sold" in the P&L sheet
# is treated as a revenue sub-line by the upload parser.
STANDARD_PNL_ITEMS = {
    "Revenue", "Total Revenue", "Cost of Goods Sold", "Gross Profit",
    "SG&A", "R&D", "D&A", "Amortization of Intangibles", "Other OpEx",
    "EBIT", "Interest Income", "Interest Expense",
    "Other Non-Operating Income / (Expense)", "EBT", "Tax", "Net Income",
}


@dataclass
class ValidationError:
    tab: str
    line_item: str
    year: int
    error_message: str


def _get(data: Dict, key: str, year: int, default=None) -> Optional[Decimal]:
    """Retrieve a value from the parsed data dict {line_item: {year: value}}."""
    item = data.get(key, {})
    val = item.get(year, default)
    if val is None:
        return None
    return Decimal(str(val))


def _get_revenue_total(pnl: Dict[str, Dict[int, Decimal]], year: int) -> Decimal:
    """
    Return the total revenue for a given year.

    Priority:
    1. "Revenue" line (standard / legacy)
    2. "Total Revenue" line (multi-stream template output)
    3. Sum of all non-standard PNL items (revenue sub-lines)
    """
    rev = _get(pnl, "Revenue", year)
    if rev is not None:
        return rev

    total_rev = _get(pnl, "Total Revenue", year)
    if total_rev is not None:
        return total_rev

    # Sum non-standard items as revenue sub-lines
    sub_total = Decimal("0")
    found_any = False
    for key, year_vals in pnl.items():
        if key not in STANDARD_PNL_ITEMS:
            val = year_vals.get(year)
            if val is not None:
                sub_total += Decimal(str(val))
                found_any = True

    return sub_total if found_any else Decimal("0")


def validate_historical_data(
    pnl: Dict[str, Dict[int, Decimal]],
    bs: Dict[str, Dict[int, Decimal]],
    cf: Dict[str, Dict[int, Decimal]],
    years: List[int],
) -> List[ValidationError]:
    errors: List[ValidationError] = []

    # Determine revenue mode once (applies to all years)
    sub_revenue_lines = [k for k in pnl if k not in STANDARD_PNL_ITEMS]
    has_sub_revenue = len(sub_revenue_lines) > 0

    for year in years:
        # --- Rule 7: No blank required fields ---
        # "Revenue" is only required when no sub-lines are defined
        pnl_required_base = [
            "Cost of Goods Sold", "Gross Profit",
            "SG&A", "R&D", "D&A", "Amortization of Intangibles", "Other OpEx",
            "EBIT", "Interest Income", "Interest Expense",
            "Other Non-Operating Income / (Expense)", "EBT", "Tax", "Net Income",
        ]
        if not has_sub_revenue:
            pnl_required_base = ["Revenue"] + pnl_required_base

        for item in pnl_required_base:
            if _get(pnl, item, year) is None:
                errors.append(ValidationError("P&L", item, year, f"'{item}' is required but missing."))

        # If sub-lines used: require at least one to have data for this year
        if has_sub_revenue:
            sub_values = [pnl[k].get(year) for k in sub_revenue_lines if pnl[k].get(year) is not None]
            if not sub_values:
                errors.append(ValidationError(
                    "P&L", "Revenue (sub-lines)", year,
                    "At least one revenue sub-line must have a value."
                ))

        bs_required = [
            "PP&E Gross", "Accumulated Depreciation", "Net PP&E",
            "Intangibles Gross", "Accumulated Amortization", "Net Intangibles", "Goodwill",
            "Inventories", "Accounts Receivable", "Prepaid Expenses & Other Current Assets",
            "Accounts Payable", "Accrued Liabilities", "Other Current Liabilities",
            "Cash & Equivalents", "Non-Operating Assets",
            "Short-Term Debt", "Long-Term Debt",
            "Share Capital", "Retained Earnings", "Other Equity (AOCI, Treasury Stock, etc.)",
        ]
        for item in bs_required:
            if _get(bs, item, year) is None:
                errors.append(ValidationError("Balance Sheet", item, year, f"'{item}' is required but missing."))

        if any(e.year == year for e in errors):
            continue

        # --- Rule 3: Gross Profit ---
        revenue = _get_revenue_total(pnl, year)

        # Cross-check "Total Revenue" vs sub-line sum
        if has_sub_revenue and "Total Revenue" in pnl:
            sub_sum = sum(
                Decimal(str(pnl[k][year]))
                for k in sub_revenue_lines
                if pnl[k].get(year) is not None
            )
            stated_total = _get(pnl, "Total Revenue", year, Decimal(0))
            if abs(sub_sum - stated_total) > TOLERANCE:
                errors.append(ValidationError(
                    "P&L", "Total Revenue", year,
                    f"Sub-lines sum to {sub_sum} but Total Revenue = {stated_total}. "
                    f"Difference: {sub_sum - stated_total}"
                ))

        cogs = _get(pnl, "Cost of Goods Sold", year, Decimal(0))
        gross_profit = _get(pnl, "Gross Profit", year, Decimal(0))
        expected_gp = revenue + cogs
        if abs(expected_gp - gross_profit) > TOLERANCE:
            errors.append(ValidationError(
                "P&L", "Gross Profit", year,
                f"Revenue ({revenue}) + COGS ({cogs}) = {expected_gp}, but Gross Profit = {gross_profit}. "
                f"Difference: {expected_gp - gross_profit}"
            ))

        # --- Rule 4: EBIT ---
        sga = _get(pnl, "SG&A", year, Decimal(0))
        rd = _get(pnl, "R&D", year, Decimal(0))
        da = _get(pnl, "D&A", year, Decimal(0))
        amort = _get(pnl, "Amortization of Intangibles", year, Decimal(0))
        other_opex = _get(pnl, "Other OpEx", year, Decimal(0))
        ebit = _get(pnl, "EBIT", year, Decimal(0))
        expected_ebit = gross_profit + sga + rd + da + amort + other_opex
        if abs(expected_ebit - ebit) > TOLERANCE:
            errors.append(ValidationError(
                "P&L", "EBIT", year,
                f"Gross Profit + OpEx = {expected_ebit}, but EBIT = {ebit}. Difference: {expected_ebit - ebit}"
            ))

        # --- Rule 5: EBT ---
        interest_income = _get(pnl, "Interest Income", year, Decimal(0))
        interest_expense = _get(pnl, "Interest Expense", year, Decimal(0))
        other_nonop = _get(pnl, "Other Non-Operating Income / (Expense)", year, Decimal(0))
        ebt = _get(pnl, "EBT", year, Decimal(0))
        expected_ebt = ebit + interest_income + interest_expense + other_nonop
        if abs(expected_ebt - ebt) > TOLERANCE:
            errors.append(ValidationError(
                "P&L", "EBT", year,
                f"EBIT ± Non-Op = {expected_ebt}, but EBT = {ebt}. Difference: {expected_ebt - ebt}"
            ))

        # --- Rule 6: Net Income ---
        tax = _get(pnl, "Tax", year, Decimal(0))
        net_income = _get(pnl, "Net Income", year, Decimal(0))
        expected_ni = ebt + tax
        if abs(expected_ni - net_income) > TOLERANCE:
            errors.append(ValidationError(
                "P&L", "Net Income", year,
                f"EBT ({ebt}) + Tax ({tax}) = {expected_ni}, but Net Income = {net_income}. "
                f"Difference: {expected_ni - net_income}"
            ))

        # --- Rule 1: Balance Sheet balance ---
        net_ppe = _get(bs, "Net PP&E", year, Decimal(0))
        net_intangibles = _get(bs, "Net Intangibles", year, Decimal(0))
        goodwill = _get(bs, "Goodwill", year, Decimal(0))
        inventories = _get(bs, "Inventories", year, Decimal(0))
        ar = _get(bs, "Accounts Receivable", year, Decimal(0))
        prepaid = _get(bs, "Prepaid Expenses & Other Current Assets", year, Decimal(0))
        cash = _get(bs, "Cash & Equivalents", year, Decimal(0))
        non_op_assets = _get(bs, "Non-Operating Assets", year, Decimal(0))

        total_assets = net_ppe + net_intangibles + goodwill + inventories + ar + prepaid + cash + non_op_assets

        ap = _get(bs, "Accounts Payable", year, Decimal(0))
        accrued = _get(bs, "Accrued Liabilities", year, Decimal(0))
        other_cl = _get(bs, "Other Current Liabilities", year, Decimal(0))
        other_lt = _get(bs, "Other Long-Term Liabilities", year, Decimal(0))
        st_debt = _get(bs, "Short-Term Debt", year, Decimal(0))
        lt_debt = _get(bs, "Long-Term Debt", year, Decimal(0))
        total_liabilities = ap + accrued + other_cl + other_lt + st_debt + lt_debt

        share_capital = _get(bs, "Share Capital", year, Decimal(0))
        retained_earnings = _get(bs, "Retained Earnings", year, Decimal(0))
        other_equity = _get(bs, "Other Equity (AOCI, Treasury Stock, etc.)", year, Decimal(0))
        total_equity = share_capital + retained_earnings + other_equity
        total_le = total_liabilities + total_equity

        if abs(total_assets + total_le) > TOLERANCE:
            errors.append(ValidationError(
                "Balance Sheet", "Total Assets vs L+E", year,
                f"Assets ({total_assets}) + Liabilities & Equity ({total_le}) should equal 0. "
                f"Sum: {total_assets + total_le}"
            ))

    # --- Rule 2: Cash reconciliation (multi-year) ---
    sorted_years = sorted(years)
    for i, year in enumerate(sorted_years[1:], 1):
        prev_year = sorted_years[i - 1]
        cash_prev = _get(bs, "Cash & Equivalents", prev_year, Decimal(0))
        cash_curr = _get(bs, "Cash & Equivalents", year, Decimal(0))
        net_change = _get(cf, "Net Change in Cash", year)
        if net_change is not None and cash_prev is not None:
            expected_cash = cash_prev + net_change
            if abs(expected_cash - cash_curr) > TOLERANCE:
                errors.append(ValidationError(
                    "Cash Flow", "Net Change in Cash", year,
                    f"Cash({prev_year}) + ΔCash = {expected_cash}, but Cash({year}) = {cash_curr}. "
                    f"Difference: {expected_cash - cash_curr}"
                ))

    return errors


def parse_historical_excel(file_bytes: bytes):
    """Parse the uploaded Excel file into structured dicts.

    Handles both:
    - Standard template: single "Revenue" row
    - Multi-stream template: "Venta Energía", "Venta Fertilizante", ... + "Total Revenue" formula

    Returns (result_dict, years, detected_revenue_sub_lines).
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    result = {"PNL": {}, "BS": {}, "CF": {}}
    detected_sub_lines: List[str] = []

    def parse_sheet(ws, key):
        headers = [cell.value for cell in ws[1]]
        year_cols = []
        years_found = []

        for col_idx, h in enumerate(headers):
            if h is not None:
                try:
                    y = int(h)
                    years_found.append(y)
                    year_cols.append(col_idx)
                except (ValueError, TypeError):
                    pass

        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = row[0]
            if not item:
                continue
            item = str(item).strip()
            data[item] = {}
            for i, year in enumerate(years_found):
                col_idx = year_cols[i]
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    try:
                        data[item][year] = Decimal(str(val))
                    except Exception:
                        pass
        result[key] = data
        return years_found

    years = []
    if "P&L" in wb.sheetnames:
        years = parse_sheet(wb["P&L"], "PNL")
    if "Balance Sheet" in wb.sheetnames:
        parse_sheet(wb["Balance Sheet"], "BS")
    if "Cash Flow" in wb.sheetnames:
        parse_sheet(wb["Cash Flow"], "CF")

    # Detect revenue sub-lines: PNL items before "Cost of Goods Sold" not in standard list
    pnl_data = result.get("PNL", {})
    for k in list(pnl_data.keys()):
        if k == "Cost of Goods Sold":
            break
        if k not in STANDARD_PNL_ITEMS:
            detected_sub_lines.append(k)

    return result, years, detected_sub_lines
