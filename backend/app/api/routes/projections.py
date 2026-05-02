import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Dict, List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import insert
from sqlalchemy.orm import Session, joinedload

from app.api.deps import get_current_user, get_project_for_write, get_project_or_404
from app.db.base import get_db
from app.models.entity import Entity
from app.models.project import (
    HistoricalData,
    NOLBalance,
    Project,
    ProjectedFinancial,
)
from app.models.user import User
from app.services.projections_runner import load_assumptions, load_historical, run_projection_engine
from app.tasks import run_projections_async

router = APIRouter(prefix="/projects", tags=["projections"])





@router.post("/{project_id}/projections/run")
def run_projection(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    project = get_project_for_write(project_id, current_user, db)

    if project.projection_years > 10:
        task = run_projections_async.delay(project_id)
        return Response(status_code=202, content=f'{{"task_id": "{task.id}", "status": "processing"}}', media_type="application/json")

    # Loop per entity: each subsidiary keeps its own historical, its own
    # assumptions, and its own projected output. The consolidated view
    # aggregates these on read; we never overwrite one entity's projections
    # with another's.
    entities = (
        db.query(Entity)
        .filter(Entity.project_id == project.id)
        .order_by(Entity.display_order)
        .all()
    )
    if not entities:
        raise HTTPException(status_code=500, detail="Project has no entity — data integrity error.")

    # NOL is project-level (single accumulator across the group). Reset once.
    db.query(NOLBalance).filter(NOLBalance.project_id == project_id).delete()

    aggregated_warnings: list[str] = []
    aggregated_proj_years: list[int] = []
    runs_persisted = 0
    skipped: list[str] = []
    nol_persisted = False

    for entity in entities:
        pnl, bs, cf, hist_years = load_historical(project_id, db, entity_id=entity.id)
        if not hist_years:
            skipped.append(f"{entity.name} (no historical)")
            continue

        assumptions = load_assumptions(project_id, db, entity_id=entity.id)
        if not assumptions:
            skipped.append(f"{entity.name} (no assumptions)")
            continue

        result, proj_years = run_projection_engine(project, pnl, bs, cf, hist_years, assumptions)

        if result.errors:
            # One entity's broken assumptions shouldn't kill the whole batch,
            # but the user must be able to see which entity failed and why.
            raise HTTPException(
                status_code=422,
                detail={
                    "error": {
                        "code": "PROJECTION_ERROR",
                        "message": f"Projection engine errors for entity '{entity.name}'",
                        "entity_id": entity.id,
                        "details": result.errors,
                    }
                },
            )

        # Wipe this entity's existing base-scenario projections, then bulk insert.
        db.query(ProjectedFinancial).filter(
            ProjectedFinancial.project_id == project_id,
            ProjectedFinancial.entity_id == entity.id,
            ProjectedFinancial.scenario_id == None,  # noqa: E711
        ).delete()

        rows: list[dict] = []
        for stmt_type, data in (("PNL", result.pnl), ("BS", result.bs), ("CF", result.cf)):
            for line_item, year_vals in data.items():
                for year, value in year_vals.items():
                    rows.append({
                        "id": str(uuid.uuid4()),
                        "project_id": project_id,
                        "entity_id": entity.id,
                        "statement_type": stmt_type,
                        "line_item": line_item,
                        "year": year,
                        "value": value,
                    })
        if rows:
            db.execute(insert(ProjectedFinancial), rows)

        # Persist NOL only from the first entity that produces it; until we
        # have a per-entity NOL model, project-level NOL = first entity's run.
        if not nol_persisted and result.nol_balances:
            nol_rows = [
                {
                    "id": str(uuid.uuid4()),
                    "project_id": project_id,
                    "year": year,
                    "nol_opening": nol["nol_opening"],
                    "nol_used": nol["nol_used"],
                    "nol_closing": nol["nol_closing"],
                }
                for year, nol in result.nol_balances.items()
            ]
            db.execute(insert(NOLBalance), nol_rows)
            nol_persisted = True

        runs_persisted += 1
        aggregated_warnings.extend(f"[{entity.name}] {w}" for w in result.warnings)
        aggregated_proj_years = proj_years  # projection horizon is project-level

    if runs_persisted == 0:
        # No entity had both historical AND assumptions — surface the same
        # 400 the legacy single-entity path used to throw, plus a hint.
        raise HTTPException(
            status_code=400,
            detail=(
                "No entity has both historical data and assumptions yet. "
                + ("Skipped: " + "; ".join(skipped) if skipped else "")
            ),
        )

    project.status = "projected"
    project.updated_at = datetime.now(timezone.utc)
    db.commit()

    return {
        "message": f"Projections complete — {runs_persisted} entity run(s)",
        "projection_years": aggregated_proj_years,
        "warnings": aggregated_warnings,
        "skipped_entities": skipped,
    }


@router.get("/{project_id}/run/status/{task_id}")
def check_projection_status(
    project_id: str,
    task_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    get_project_or_404(project_id, current_user, db)
    res = run_projections_async.AsyncResult(task_id)
    if res.ready():
        res_data = res.get()
        if res_data.get("status") == "error":
            raise HTTPException(422, detail={"error": {"code": "PROJECTION_ERROR", "message": res_data.get("detail", "Error"), "details": res_data.get("errors", [])}})
        return {"status": "completed"}
    return {"status": "processing"}


@router.get("/{project_id}/projections")
def get_projections(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    get_project_or_404(project_id, current_user, db)

    # Include historical data for context
    hist_records = db.query(HistoricalData).filter(HistoricalData.project_id == project_id).all()
    result = {"PNL": {}, "BS": {}, "CF": {}}
    historical_years = set()
    projected_years = set()

    for r in hist_records:
        result[r.statement_type].setdefault(r.line_item, {})[r.year] = str(r.value)
        historical_years.add(r.year)

    # Overlay projected data (projected years take precedence)
    proj_records = db.query(ProjectedFinancial).filter(ProjectedFinancial.project_id == project_id).all()
    for r in proj_records:
        result[r.statement_type].setdefault(r.line_item, {})[r.year] = str(r.value)
        projected_years.add(r.year)

    result["historical_years"] = sorted(historical_years)
    result["projected_years"] = sorted(projected_years)

    return result


@router.get("/{project_id}/projections/export")
def export_projections(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export projected financials to Excel."""
    project = get_project_or_404(project_id, current_user, db)
    pnl, bs, cf, hist_years = load_historical(project_id, db)
    proj_records = db.query(ProjectedFinancial).filter(ProjectedFinancial.project_id == project_id).all()

    if not proj_records:
        raise HTTPException(400, "No projections available. Run the projection engine first.")

    # Build projected dicts
    proj_pnl, proj_bs, proj_cf = {}, {}, {}
    proj_years = set()
    for r in proj_records:
        proj_years.add(r.year)
        if r.statement_type == "PNL":
            proj_pnl.setdefault(r.line_item, {})[r.year] = Decimal(str(r.value))
        elif r.statement_type == "BS":
            proj_bs.setdefault(r.line_item, {})[r.year] = Decimal(str(r.value))
        elif r.statement_type == "CF":
            proj_cf.setdefault(r.line_item, {})[r.year] = Decimal(str(r.value))

    proj_years = sorted(proj_years)
    all_years = hist_years + proj_years

    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    PROJ_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    def write_export_sheet(ws, line_items, hist_data, proj_data):
        headers = ["Line Item"] + [str(y) for y in all_years]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            if col > 1 + len(hist_years):
                cell.fill = PROJ_FILL

        for row_idx, li in enumerate(line_items, 2):
            ws.cell(row=row_idx, column=1, value=li)
            for col_idx, year in enumerate(all_years, 2):
                if year in hist_years:
                    val = hist_data.get(li, {}).get(year)
                else:
                    val = proj_data.get(li, {}).get(year)
                cell = ws.cell(row=row_idx, column=col_idx, value=float(val) if val is not None else None)
                if year in proj_years:
                    cell.fill = PROJ_FILL

        ws.column_dimensions["A"].width = 45
        for i in range(len(all_years)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 14

    from app.services.template_generator import BS_ITEMS, CF_ITEMS, PNL_ITEMS

    ws_pnl = wb.active
    ws_pnl.title = "P&L"
    write_export_sheet(ws_pnl, [item for item, _ in PNL_ITEMS], pnl, proj_pnl)

    ws_bs = wb.create_sheet("Balance Sheet")
    write_export_sheet(ws_bs, [item for item, _, _ in BS_ITEMS], bs, proj_bs)

    ws_cf = wb.create_sheet("Cash Flow")
    write_export_sheet(ws_cf, [item for item, _ in CF_ITEMS], cf, proj_cf)

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={project.name}_projections.xlsx"},
    )
